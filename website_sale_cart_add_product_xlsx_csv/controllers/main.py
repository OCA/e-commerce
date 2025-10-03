# Copyright 2025 Alberto Martínez <alberto.martinez@sygel.es>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import base64
import csv
import io
import math

from openpyxl import load_workbook

from odoo import _, http
from odoo.exceptions import UserError
from odoo.http import content_disposition, request

from odoo.addons.website_sale.controllers.main import WebsiteSale


class WebsiteSaleAddProductXlsxCsv(WebsiteSale):
    def _parse_file(self, file):
        headers = False
        data = False
        if file.filename.endswith(".csv"):
            file.stream.seek(0)
            text_stream = io.StringIO(file.read().decode("utf-8"))
            reader = csv.DictReader(text_stream)
            headers = reader.fieldnames
            rows = [row for row in reader if any(row.values())]
            data = enumerate(rows, start=2)
            data = rows

        elif file.filename.endswith(".xlsx"):
            file.stream.seek(0)
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = [str(h).strip() for h in rows[0]]
            rows = rows[1:]
            data = [dict(zip(headers, row, strict=True)) for row in rows if any(row)]

        return headers, data

    def _can_be_imported(self, product, qty, sale):
        can_be_imported = True
        error_reason = ""
        try:
            qty = float(qty)
        except ValueError:
            can_be_imported = False
            error_reason = _("The quantity is not a number")
        if not isinstance(qty, int | float) or math.isnan(qty) or qty <= 0:
            can_be_imported = False
            error_reason = _("The quantity should be a positive number")
        if len(product) != 1:
            can_be_imported = False
            error_reason = _("The product was not found")
        elif not product.active:
            can_be_imported = False
            error_reason = _("The product is archived")
        elif not product.sale_ok:
            can_be_imported = False
            error_reason = _("Product can not be sold")
        elif not product.website_published:
            can_be_imported = False
            error_reason = _("Product is not published on the website")
        elif (
            not product.allow_out_of_stock_order
            and qty > product.with_context(warehouse=sale.warehouse_id.id).free_qty
        ):
            can_be_imported = False
            qty = product.with_context(warehouse=sale.warehouse_id.id).free_qty
            error_reason = _("Product max saleable qty is {}").format(qty)
        return can_be_imported, error_reason

    def _check_file_size(self, file):
        file_limit_mb = request.website.cart_import_button_file_limit
        file_size_bytes = file.stream.seek(0, 2)
        file_size_ok = file_size_bytes / 1024**2 <= file_limit_mb
        file.stream.seek(0)
        return file_size_ok

    def import_file(self, sale_order, file):
        import_status = "sucess"
        import_msg = ""
        headers = data = False
        failed_products = []
        if not self._check_file_size(file):
            import_status = "error"
            import_msg = _(
                "The file size is greater than the maximum allowed, {} MB"
            ).format(request.website.cart_import_button_file_limit)

        if import_status != "error" and not (
            file.filename.endswith(".csv") or file.filename.endswith(".xlsx")
        ):
            import_status = "error"
            import_msg = _("Incorrect file format, it must me a .xlsx or a .csv")
        else:
            headers, data = self._parse_file(file)

        if import_status != "error" and headers != ["default_code", "product_uom_qty"]:
            import_status = "error"
            import_msg = _(
                "Incorrect file format, "
                "the columns should be 'default_code' and 'product_uom_qty'"
            )

        if import_status != "error":
            for index, row in enumerate(data, start=2):
                default_code = str(row["default_code"]).strip()
                qty = row["product_uom_qty"]
                product = (
                    request.env["product.product"]
                    .with_context(active_test=False)
                    .search([("default_code", "=", default_code)])
                )
                can_be_imported, warn_msg = self._can_be_imported(
                    product, qty, sale_order
                )
                if not can_be_imported:
                    import_status = "warn"
                    failed_products.append(
                        f"Line {index}. {row['default_code']}: {warn_msg}"
                    )
                else:
                    try:
                        sale_order._cart_update(
                            product_id=product.id, set_qty=float(qty)
                        )
                    except UserError as e:
                        import_status = "warn"
                        failed_products.append(
                            f"Line {index}. {row['default_code']}: {str(e)}"
                        )

        return import_status, import_msg, failed_products

    @http.route("/shop/cart", type="http", auth="public", website=True, sitemap=False)
    def cart(self, access_token=None, revive="", **post):
        file = post.get("cart_file")
        if file:
            sale_order = request.website.sale_get_order(force_create=True)
            import_status, import_msg, failed_products = self.import_file(
                sale_order, file
            )
            post.update(
                {
                    "import_status": import_status,
                    "import_msg": import_msg,
                    "failed_products": failed_products,
                }
            )
        return super().cart(access_token, revive, **post)

    def _cart_values(self, **post):
        res = super()._cart_values(**post)
        if "import_status" in post:
            res["import_status"] = post.get("import_status")
        if "import_msg" in post:
            res["import_msg"] = post.get("import_msg")
        if "failed_products" in post:
            res["failed_products"] = post.get("failed_products")
        return res

    @http.route("/shop/cart/import/example", auth="public")
    def cart_import_example(self):
        attachment = request.env.ref(
            "website_sale_cart_add_product_xlsx_csv.cart_import_example"
        ).sudo()
        filecontent = base64.b64decode(attachment.datas)
        filename = f"{attachment.name}.xlsx"

        return request.make_response(
            filecontent,
            [
                ("Content-Type", attachment.mimetype),
                ("Content-Disposition", content_disposition(filename)),
            ],
        )
